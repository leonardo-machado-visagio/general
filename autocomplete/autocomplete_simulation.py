#!/usr/bin/env python3
"""Simulação de autocomplete para aula de LLMs.

Faz N chamadas paralelas com prefill para cada frase incompleta e
gera análise da distribuição da próxima palavra (frequência, palavras
únicas, entropia de Shannon e massa da cauda).
"""

import asyncio
import json
import math
import os
import random
import re
import sys
import time
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

OUTPUT_DIR = Path(__file__).resolve().parent / "output"

try:
    from anthropic import AsyncAnthropic, APIStatusError, RateLimitError
except ImportError:
    sys.stderr.write("ERRO: pacote `anthropic` não instalado.\n")
    sys.stderr.write("Rode: pip install -r requirements.txt\n")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    sys.stderr.write("ERRO: pacote `openpyxl` não instalado.\n")
    sys.stderr.write("Rode: pip install -r requirements.txt\n")
    sys.exit(1)


# === CONFIG ===

QUESTIONS = {
    "Q1_pouco_contexto": "A ação foi",
    "Q2_contexto_medio": "A ação de despejo foi",
    "Q3_muito_contexto": "Após o inquilino purgar a mora dentro do prazo legal, a ação de despejo foi",
}

N_PER_QUESTION = int(os.environ.get("N_PER_QUESTION", "10000"))
CONCURRENCY = 50
MAX_RETRIES = 5
MAX_TOKENS = 15
TEMPERATURE = 1.0
STOP_SEQUENCES = []

# Bound retry behavior so a stuck call cannot block the whole gather indefinitely.
CALL_TIMEOUT_S = 60
BACKOFF_CAP_S = 15

MODEL = "claude-haiku-4-5-20251001"
MODEL_DISPLAY_NAME = "Claude Haiku 4.5"

# Pricing per million tokens, in USD. Used to compute exact cost from response.usage.
PRICING = {
    "claude-haiku-4-5-20251001": {"input": 1.00, "output": 5.00},
    "claude-opus-4-7": {"input": 15.00, "output": 75.00},
    "claude-opus-4-6": {"input": 15.00, "output": 75.00},
}

WORD_REGEX = re.compile(r"[a-zA-ZáéíóúâêôãõàçüÁÉÍÓÚÂÊÔÃÕÀÇÜ]+")

USER_INSTRUCTION = "Continue a frase com a continuação mais natural. Apenas continue, sem comentários."


# === API ===

def get_api_key():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        sys.stderr.write(
            "ERRO: variável de ambiente ANTHROPIC_API_KEY não encontrada.\n"
            "Configure antes de rodar:\n\n"
            "    export ANTHROPIC_API_KEY=sk-ant-...\n\n"
        )
        sys.exit(1)
    return api_key


async def detect_model(client):
    for attempt in range(6):
        try:
            await asyncio.wait_for(
                client.messages.create(
                    model=MODEL,
                    max_tokens=1,
                    messages=[{"role": "user", "content": "oi"}],
                ),
                timeout=CALL_TIMEOUT_S,
            )
            print(f"Modelo selecionado: {MODEL}")
            return MODEL
        except (APIStatusError, asyncio.TimeoutError) as e:
            status = getattr(e, "status_code", None)
            if attempt < 5:
                backoff = min(2 ** attempt, BACKOFF_CAP_S)
                print(f"  probe falhou ({status or type(e).__name__}); retry em {backoff}s...")
                await asyncio.sleep(backoff)
                continue
            print(f"  desistindo do probe após {attempt + 1} tentativas")
            break
    sys.stderr.write(f"ERRO: modelo {MODEL} não respondeu ao probe.\n")
    sys.exit(1)


# === PROGRESS & USAGE ===

class Progress:
    def __init__(self, total, label=""):
        self.total = total
        self.label = label
        self.done = 0
        self.lock = asyncio.Lock()
        self.start = time.time()

    async def tick(self):
        async with self.lock:
            self.done += 1
            if self.done % 100 == 0 or self.done == self.total:
                elapsed = time.time() - self.start
                rate = self.done / elapsed if elapsed > 0 else 0
                remaining = self.total - self.done
                eta = remaining / rate if rate > 0 else 0
                tag = f"[{self.label}] " if self.label else ""
                print(
                    f"  {tag}{self.done:>5}/{self.total}  "
                    f"({100 * self.done / self.total:5.1f}%)  "
                    f"{rate:5.1f} req/s  ETA {eta:5.0f}s"
                )


class UsageTracker:
    def __init__(self):
        self.input_tokens = 0
        self.output_tokens = 0

    def add(self, response):
        usage = getattr(response, "usage", None)
        if usage is None:
            return
        self.input_tokens += getattr(usage, "input_tokens", 0) or 0
        self.output_tokens += getattr(usage, "output_tokens", 0) or 0

    def cost_usd(self, model):
        prices = PRICING.get(model)
        if not prices:
            return None
        return (
            self.input_tokens / 1_000_000 * prices["input"]
            + self.output_tokens / 1_000_000 * prices["output"]
        )


# === CALLS ===

def _backoff(attempt):
    return min(2 ** attempt, BACKOFF_CAP_S) + random.uniform(0, 1)


async def call_one(client, model, prefill, semaphore, progress, usage):
    async with semaphore:
        last_error = None
        for attempt in range(MAX_RETRIES):
            try:
                kwargs = dict(
                    model=model,
                    max_tokens=MAX_TOKENS,
                    temperature=TEMPERATURE,
                    messages=[
                        {"role": "user", "content": USER_INSTRUCTION},
                        {"role": "assistant", "content": prefill},
                    ],
                )
                if STOP_SEQUENCES:
                    kwargs["stop_sequences"] = STOP_SEQUENCES
                response = await asyncio.wait_for(
                    client.messages.create(**kwargs),
                    timeout=CALL_TIMEOUT_S,
                )
                usage.add(response)
                text = response.content[0].text if response.content else ""
                await progress.tick()
                return text
            except asyncio.TimeoutError:
                last_error = TimeoutError(f"call exceeded {CALL_TIMEOUT_S}s")
                if attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(_backoff(attempt))
                    continue
                break
            except RateLimitError as e:
                last_error = e
                if attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(_backoff(attempt))
                    continue
                break
            except APIStatusError as e:
                last_error = e
                status = getattr(e, "status_code", None)
                if status and (status == 429 or status >= 500) and attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(_backoff(attempt))
                    continue
                break
            except Exception as e:
                last_error = e
                if attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(_backoff(attempt))
                    continue
                break
        await progress.tick()
        return f"<ERROR: {type(last_error).__name__}: {last_error}>"


# === ANALYSIS ===

def extract_first_word(text):
    if not text or text.startswith("<ERROR"):
        return None
    m = WORD_REGEX.search(text)
    return m.group(0).lower() if m else None


def entropy_bits(counter, total_valid):
    """Shannon entropy of the next-word distribution, in bits."""
    if total_valid == 0:
        return 0.0
    return -sum(
        (c / total_valid) * math.log2(c / total_valid)
        for c in counter.values()
        if c > 0
    )


def tail_mass_pct(counter, total_valid, top_k=10):
    """Percentage of probability mass that lives outside the top-k words."""
    if total_valid == 0:
        return 0.0
    top_count = sum(c for _, c in counter.most_common(top_k))
    return 100 * (1 - top_count / total_valid)


def build_distributions(results):
    distributions = {}
    for q_key, data in results.items():
        counter = Counter()
        errors = 0
        for response in data["responses"]:
            if response.startswith("<ERROR"):
                errors += 1
                continue
            word = extract_first_word(response)
            if word:
                counter[word] += 1
        total_valid = sum(counter.values())
        distributions[q_key] = {
            "counter": counter,
            "total_valid": total_valid,
            "errors": errors,
            "unique_words": len(counter),
            "entropy_bits": entropy_bits(counter, total_valid),
            "tail_mass_pct": tail_mass_pct(counter, total_valid, 10),
        }
    return distributions


def derive_label(key):
    """'Q1_pouco_contexto' -> 'Q1 — Pouco contexto'."""
    parts = key.split("_", 1)
    if len(parts) < 2:
        return key
    prefix, rest = parts
    return f"{prefix} — {rest.replace('_', ' ').capitalize()}"


# === REPORTS ===

def write_excel(distributions, metadata, path):
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")

    def header_row(ws, row_idx):
        for c in ws[row_idx]:
            c.font = header_font
            c.fill = header_fill

    # Metadata sheet
    meta = wb.create_sheet("Metadata")
    meta.append(["chave", "valor"])
    header_row(meta, 1)
    meta.append(["timestamp", metadata["timestamp"]])
    meta.append(["model", metadata["model"]])
    meta.append(["n_per_question", metadata["n_per_question"]])
    meta.append(["concurrency", metadata["concurrency"]])
    meta.append(["temperature", metadata["temperature"]])
    meta.append(["max_tokens", metadata["max_tokens"]])
    meta.append(["stop_sequences", json.dumps(metadata["stop_sequences"])])
    meta.append(["duration_s", round(metadata["duration_s"], 1)])
    meta.append(["input_tokens", metadata["usage"]["input_tokens"]])
    meta.append(["output_tokens", metadata["usage"]["output_tokens"]])
    cost = metadata["usage"].get("cost_usd")
    meta.append(["cost_usd", f"{cost:.4f}" if cost is not None else "n/a"])
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 42

    # Resumo with the new metrics on top, then top-10 table, then prompts.
    resumo = wb.create_sheet("Resumo")
    resumo.append(["métrica"] + list(QUESTIONS))
    header_row(resumo, 1)

    def metric_row(label, fn):
        resumo.append([label] + [fn(distributions[q]) for q in QUESTIONS])

    metric_row("respostas válidas", lambda d: d["total_valid"])
    metric_row("erros", lambda d: d["errors"])
    metric_row("palavras únicas", lambda d: d["unique_words"])
    metric_row("entropia (bits)", lambda d: round(d["entropy_bits"], 2))
    metric_row("massa fora do top 10 (%)", lambda d: round(d["tail_mass_pct"], 1))

    resumo.append([])
    resumo.append(["rank"] + [f"{q} — top palavra (%)" for q in QUESTIONS])
    header_row(resumo, resumo.max_row)
    for i in range(10):
        row = [i + 1]
        for q_key in QUESTIONS:
            counter = distributions[q_key]["counter"]
            top = counter.most_common(10)
            if i < len(top):
                w, c = top[i]
                total = distributions[q_key]["total_valid"]
                pct = 100 * c / total if total else 0
                row.append(f"{w} ({pct:.1f}%)")
            else:
                row.append("")
        resumo.append(row)

    resumo.append([])
    resumo.append(["frases (prefill):"])
    resumo.cell(row=resumo.max_row, column=1).font = Font(bold=True)
    for q_key, prefill in QUESTIONS.items():
        resumo.append([q_key, prefill])

    for col_idx in range(1, len(QUESTIONS) + 2):
        col_letter = resumo.cell(row=1, column=col_idx).column_letter
        resumo.column_dimensions[col_letter].width = 38

    # Per-question sheets
    for q_key in QUESTIONS:
        ws = wb.create_sheet(q_key)
        ws.append(["palavra", "contagem", "percentual"])
        header_row(ws, 1)
        counter = distributions[q_key]["counter"]
        total = distributions[q_key]["total_valid"]
        for word, count in counter.most_common():
            pct = 100 * count / total if total else 0
            ws.append([word, count, round(pct, 2)])
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 14

    wb.save(path)


def build_html(distributions, metadata):
    chart_data = {}
    for q_key, info in distributions.items():
        counter = info["counter"]
        total = info["total_valid"]
        top10 = counter.most_common(10)
        chart_data[q_key] = {
            "label": derive_label(q_key),
            "prompt": QUESTIONS[q_key],
            "labels": [w for w, _ in top10],
            "counts": [c for _, c in top10],
            "percentages": [round(100 * c / total, 2) if total else 0 for _, c in top10],
            "total_valid": total,
            "errors": info["errors"],
            "unique_words": info["unique_words"],
            "entropy_bits": round(info["entropy_bits"], 2),
            "tail_mass_pct": round(info["tail_mass_pct"], 1),
        }
    payload = {
        "questions": chart_data,
        "order": list(QUESTIONS),
        "model_display": metadata["model_display"],
        "n_per_question": metadata["n_per_question"],
        "temperature": metadata["temperature"],
    }
    data_json = json.dumps(payload, ensure_ascii=False)
    return HTML_TEMPLATE.replace("__PAYLOAD__", data_json)


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<title>Distribuição de respostas — autocomplete LLM</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  * { box-sizing: border-box; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    max-width: 1500px;
    margin: 0 auto;
    padding: 24px;
    background: #fafafa;
    color: #1f2937;
  }
  h1 { color: #111827; margin-bottom: 4px; font-size: 26px; }
  .subtitle { color: #6b7280; margin-bottom: 28px; font-size: 14px; }
  .charts {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 20px;
    align-items: stretch;
  }
  .chart-container {
    background: white;
    border-radius: 10px;
    padding: 20px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08);
    display: flex;
    flex-direction: column;
  }
  .chart-container h2 {
    font-size: 12px;
    color: #6b7280;
    margin: 0 0 10px 0;
    font-weight: 600;
    letter-spacing: 0.8px;
    text-transform: uppercase;
  }
  .prompt {
    color: #111827;
    margin: 0 0 14px 0;
    padding: 16px 18px;
    background: #eef2ff;
    border-left: 5px solid #4f46e5;
    border-radius: 6px;
    font-size: 19px;
    font-weight: 600;
    line-height: 1.35;
    min-height: 130px;
    display: flex;
    align-items: center;
  }
  .prompt .ellipsis { color: #6366f1; font-weight: 700; }
  .meta {
    font-size: 12px;
    color: #9ca3af;
    margin-bottom: 10px;
    display: flex;
    flex-wrap: wrap;
    gap: 6px 10px;
  }
  .meta .stat { white-space: nowrap; }
  .meta .stat strong { color: #4f46e5; font-weight: 600; }
  .chart-wrapper { height: 380px; position: relative; }
  footer {
    margin-top: 32px;
    padding-top: 16px;
    border-top: 1px solid #e5e7eb;
    font-size: 12px;
    color: #9ca3af;
  }
  @media (max-width: 1100px) {
    .charts { grid-template-columns: 1fr; }
  }
</style>
</head>
<body>
<h1 id="page-title"></h1>
<p class="subtitle" id="page-subtitle"></p>
<div class="charts" id="charts"></div>
<footer>
  Gerado por <code>autocomplete_simulation.py</code>.
</footer>
<script>
const PAYLOAD = __PAYLOAD__;

document.getElementById("page-title").textContent =
  "Distribuição de respostas do " + PAYLOAD.model_display + " — efeito do contexto";
document.getElementById("page-subtitle").textContent =
  PAYLOAD.n_per_question + " chamadas por prompt · temperatura " + PAYLOAD.temperature +
  " · prefill como mensagem do assistant · primeira palavra extraída · top 10 mostrado.";

function statSpan(label, value) {
  return '<span class="stat"><strong>' + value + '</strong> ' + label + '</span>';
}

function renderQuestion(container, key) {
  const d = PAYLOAD.questions[key];

  const h2 = document.createElement("h2");
  h2.textContent = d.label;

  const prompt = document.createElement("div");
  prompt.className = "prompt";
  prompt.innerHTML = '"' + d.prompt + '<span class="ellipsis">…</span>"';

  const meta = document.createElement("div");
  meta.className = "meta";
  let metaHtml = statSpan("respostas", d.total_valid)
    + statSpan("únicas", d.unique_words)
    + statSpan("bits de entropia", d.entropy_bits)
    + statSpan("% fora do top 10", d.tail_mass_pct);
  if (d.errors > 0) {
    metaHtml += statSpan("erros", d.errors);
  }
  meta.innerHTML = metaHtml;

  const wrapper = document.createElement("div");
  wrapper.className = "chart-wrapper";
  const canvas = document.createElement("canvas");
  wrapper.appendChild(canvas);

  container.appendChild(h2);
  container.appendChild(prompt);
  container.appendChild(meta);
  container.appendChild(wrapper);

  new Chart(canvas, {
    type: 'bar',
    data: {
      labels: d.labels,
      datasets: [{
        label: '%',
        data: d.percentages,
        backgroundColor: 'rgba(79, 70, 229, 0.85)',
        borderColor: 'rgba(79, 70, 229, 1)',
        borderWidth: 1,
      }],
    },
    options: {
      indexAxis: 'y',
      responsive: true,
      maintainAspectRatio: false,
      plugins: {
        legend: { display: false },
        tooltip: {
          callbacks: {
            label: function(c) {
              return c.parsed.x.toFixed(1) + '%  (' + d.counts[c.dataIndex] + ' de ' + d.total_valid + ')';
            }
          }
        }
      },
      scales: {
        x: {
          title: { display: true, text: '% das respostas' },
          ticks: { callback: function(v) { return v + '%'; } }
        },
        y: {
          ticks: { font: { size: 13 } }
        }
      }
    }
  });
}

const chartsEl = document.getElementById("charts");
PAYLOAD.order.forEach(key => {
  const c = document.createElement("div");
  c.className = "chart-container";
  chartsEl.appendChild(c);
  renderQuestion(c, key);
});
</script>
</body>
</html>
"""


# === MAIN ===

async def main():
    print("=" * 70)
    print(f"Simulação de autocomplete — {MODEL_DISPLAY_NAME}")
    print("=" * 70)
    api_key = get_api_key()
    client = AsyncAnthropic(api_key=api_key)

    model = await detect_model(client)

    total = len(QUESTIONS) * N_PER_QUESTION
    print(f"\nTotal de chamadas: {total} ({N_PER_QUESTION} x {len(QUESTIONS)} perguntas)")
    print(
        f"Paralelismo: {CONCURRENCY} | Temperatura: {TEMPERATURE} | "
        f"max_tokens: {MAX_TOKENS} | stop_sequences: {STOP_SEQUENCES!r}"
    )
    print(f"Timeout por call: {CALL_TIMEOUT_S}s | retries até {MAX_RETRIES}\n")

    semaphore = asyncio.Semaphore(CONCURRENCY)
    usage = UsageTracker()
    t0 = time.time()

    results = {}
    for q_key, prefill in QUESTIONS.items():
        print(f"\n[{q_key}] prefill: {prefill!r}")
        progress = Progress(N_PER_QUESTION, label=q_key)
        tasks = [
            call_one(client, model, prefill, semaphore, progress, usage)
            for _ in range(N_PER_QUESTION)
        ]
        responses = await asyncio.gather(*tasks)
        results[q_key] = {"prompt": prefill, "responses": responses}

    duration = time.time() - t0
    cost = usage.cost_usd(model)
    print(f"\nConcluído em {duration:.1f}s")
    print(
        f"Uso real: {usage.input_tokens:,} input + {usage.output_tokens:,} output tokens"
    )
    if cost is not None:
        print(f"Custo real (preço de tabela): US$ {cost:.3f}\n")
    else:
        print(f"(modelo {model!r} não está na tabela de preços PRICING)\n")

    metadata = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "model": model,
        "model_display": MODEL_DISPLAY_NAME,
        "n_per_question": N_PER_QUESTION,
        "concurrency": CONCURRENCY,
        "temperature": TEMPERATURE,
        "max_tokens": MAX_TOKENS,
        "stop_sequences": list(STOP_SEQUENCES),
        "duration_s": duration,
        "usage": {
            "input_tokens": usage.input_tokens,
            "output_tokens": usage.output_tokens,
            "cost_usd": cost,
        },
    }

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    raw_path = OUTPUT_DIR / "raw_results.json"
    xlsx_path = OUTPUT_DIR / "distribuicao.xlsx"
    html_path = OUTPUT_DIR / "distribuicao.html"

    with open(raw_path, "w", encoding="utf-8") as f:
        json.dump({"metadata": metadata, "results": results}, f, ensure_ascii=False, indent=2)
    print(f"Salvo: {raw_path}")

    distributions = build_distributions(results)

    write_excel(distributions, metadata, xlsx_path)
    print(f"Salvo: {xlsx_path}")

    html = build_html(distributions, metadata)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Salvo: {html_path}")

    print("\n" + "=" * 70)
    print("TOP 5 por pergunta")
    print("=" * 70)
    for q_key, info in distributions.items():
        prefill = QUESTIONS[q_key]
        counter = info["counter"]
        total_valid = info["total_valid"]
        print(f"\n[{q_key}]  {prefill!r}")
        print(
            f"  válidas: {total_valid} · erros: {info['errors']} · "
            f"únicas: {info['unique_words']} · "
            f"entropia: {info['entropy_bits']:.2f} bits · "
            f"cauda: {info['tail_mass_pct']:.1f}%"
        )
        for word, count in counter.most_common(5):
            pct = 100 * count / total_valid if total_valid else 0
            bar = "█" * int(pct / 2)
            print(f"  {word:<25} {count:>4}  ({pct:5.1f}%) {bar}")
    print()


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        sys.stderr.write("\nInterrompido pelo usuário.\n")
        sys.exit(130)
