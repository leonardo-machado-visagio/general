#!/usr/bin/env python3
"""Simulação de autocomplete do Claude Opus para aula de LLMs para advogados.

Faz 1000 chamadas paralelas com prefill para cada uma de 3 frases incompletas
e gera análise da distribuição da próxima palavra.
"""

import asyncio
import json
import os
import random
import re
import sys
import time
from collections import Counter

try:
    from anthropic import AsyncAnthropic, APIStatusError, RateLimitError
except ImportError:
    sys.stderr.write("ERRO: pacote `anthropic` não instalado.\n")
    sys.stderr.write("Rode: pip install anthropic openpyxl\n")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    sys.stderr.write("ERRO: pacote `openpyxl` não instalado.\n")
    sys.stderr.write("Rode: pip install anthropic openpyxl\n")
    sys.exit(1)


QUESTIONS = {
    "Q1_pouco_contexto": "A ação foi",
    "Q2_contexto_medio": "A ação de despejo foi",
    "Q3_muito_contexto": "Após o inquilino purgar a mora dentro do prazo legal, a ação de despejo foi",
}

N_PER_QUESTION = 1000
CONCURRENCY = 50
MAX_RETRIES = 5
MAX_TOKENS = 5
TEMPERATURE = 1.0

PRIMARY_MODEL = "claude-haiku-4-5-20251001"
FALLBACK_MODEL = "claude-haiku-4-5-20251001"
MODEL_DISPLAY_NAME = "Claude Haiku 4.5"

WORD_REGEX = re.compile(r"[a-zA-ZáéíóúâêôãõàçüÁÉÍÓÚÂÊÔÃÕÀÇÜ]+")

USER_INSTRUCTION = "Continue a frase com a continuação mais natural. Apenas continue, sem comentários."


def get_api_key():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        sys.stderr.write(
            "ERRO: variável de ambiente ANTHROPIC_API_KEY não encontrada.\n"
            "Configure antes de rodar:\n\n"
            "    export ANTHROPIC_API_KEY=sk-ant-...\n\n"
        )
        sys.exit(1)
    return api_key


async def detect_model(client):
    for model in (PRIMARY_MODEL, FALLBACK_MODEL):
        for attempt in range(6):
            try:
                await client.messages.create(
                    model=model,
                    max_tokens=1,
                    messages=[{"role": "user", "content": "oi"}],
                )
                print(f"Modelo selecionado: {model}")
                return model
            except APIStatusError as e:
                status = getattr(e, "status_code", None)
                msg = str(e).lower()
                if status == 404 or "not_found" in msg or "model_not_found" in msg:
                    print(f"Modelo {model} indisponível ({status}), tentando fallback...")
                    break
                if attempt < 5:
                    backoff = min(2 ** attempt, 30)
                    print(f"  transient {status} no probe; retry em {backoff}s...")
                    await asyncio.sleep(backoff)
                    continue
                print(f"  desistindo do probe ({status}) após {attempt + 1} tentativas")
                break
    sys.stderr.write("ERRO: nem o modelo primário nem o fallback responderam ao probe.\n")
    sys.exit(1)


class Progress:
    def __init__(self, total):
        self.total = total
        self.done = 0
        self.lock = asyncio.Lock()
        self.start = time.time()

    async def tick(self):
        async with self.lock:
            self.done += 1
            if self.done % 100 == 0 or self.done == self.total:
                elapsed = time.time() - self.start
                rate = self.done / elapsed if elapsed > 0 else 0
                remaining = self.total - self.done
                eta = remaining / rate if rate > 0 else 0
                print(
                    f"  {self.done:>5}/{self.total}  "
                    f"({100 * self.done / self.total:5.1f}%)  "
                    f"{rate:5.1f} req/s  ETA {eta:5.0f}s"
                )


def _backoff(attempt):
    return min(2 ** attempt, 30) + random.uniform(0, 1)


async def call_one(client, model, prefill, semaphore, progress):
    async with semaphore:
        last_error = None
        for attempt in range(MAX_RETRIES):
            try:
                response = await client.messages.create(
                    model=model,
                    max_tokens=MAX_TOKENS,
                    temperature=TEMPERATURE,
                    messages=[
                        {"role": "user", "content": USER_INSTRUCTION},
                        {"role": "assistant", "content": prefill},
                    ],
                )
                text = response.content[0].text if response.content else ""
                await progress.tick()
                return text
            except RateLimitError as e:
                last_error = e
                if attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(_backoff(attempt))
                    continue
                break
            except APIStatusError as e:
                last_error = e
                status = getattr(e, "status_code", None)
                if status and (status == 429 or status >= 500) and attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(_backoff(attempt))
                    continue
                break
            except Exception as e:
                last_error = e
                if attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(_backoff(attempt))
                    continue
                break
        await progress.tick()
        return f"<ERROR: {type(last_error).__name__}: {last_error}>"


def extract_first_word(text):
    if not text or text.startswith("<ERROR"):
        return None
    m = WORD_REGEX.search(text)
    return m.group(0).lower() if m else None


def build_distributions(results):
    distributions = {}
    for q_key, data in results.items():
        counter = Counter()
        errors = 0
        for response in data["responses"]:
            if response.startswith("<ERROR"):
                errors += 1
                continue
            word = extract_first_word(response)
            if word:
                counter[word] += 1
        distributions[q_key] = {
            "counter": counter,
            "total_valid": sum(counter.values()),
            "errors": errors,
        }
    return distributions


def write_excel(distributions, path):
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")

    # Resumo first
    resumo = wb.create_sheet("Resumo")
    resumo.append(["Rank"] + [f"{k} — top palavra (%)" for k in QUESTIONS.keys()])
    for cell in resumo[1]:
        cell.font = header_font
        cell.fill = header_fill
    for i in range(10):
        row = [i + 1]
        for q_key in QUESTIONS.keys():
            counter = distributions[q_key]["counter"]
            top = counter.most_common(10)
            if i < len(top):
                word, count = top[i]
                total = distributions[q_key]["total_valid"]
                pct = 100 * count / total if total else 0
                row.append(f"{word} ({pct:.1f}%)")
            else:
                row.append("")
        resumo.append(row)
    for col in "ABCD":
        resumo.column_dimensions[col].width = 35

    # Frase usada
    resumo.append([])
    resumo.append(["Frases (prefill):"])
    resumo.cell(row=resumo.max_row, column=1).font = Font(bold=True)
    for q_key, prefill in QUESTIONS.items():
        resumo.append([q_key, prefill])

    # Por pergunta
    for q_key in QUESTIONS.keys():
        ws = wb.create_sheet(q_key)
        ws.append(["palavra", "contagem", "percentual"])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        counter = distributions[q_key]["counter"]
        total = distributions[q_key]["total_valid"]
        for word, count in counter.most_common():
            pct = 100 * count / total if total else 0
            ws.append([word, count, round(pct, 2)])
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 14

    wb.save(path)


def build_html(distributions):
    chart_data = {}
    for q_key, info in distributions.items():
        counter = info["counter"]
        total = info["total_valid"]
        top10 = counter.most_common(10)
        chart_data[q_key] = {
            "prompt": QUESTIONS[q_key],
            "labels": [w for w, _ in top10],
            "counts": [c for _, c in top10],
            "percentages": [round(100 * c / total, 2) if total else 0 for _, c in top10],
            "total_valid": total,
            "errors": info["errors"],
        }
    data_json = json.dumps(chart_data, ensure_ascii=False)

    return """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<title>Distribuição de respostas do Claude Haiku 4.5</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  * { box-sizing: border-box; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    max-width: 1500px;
    margin: 0 auto;
    padding: 24px;
    background: #fafafa;
    color: #1f2937;
  }
  h1 { color: #111827; margin-bottom: 4px; font-size: 24px; }
  .subtitle { color: #6b7280; margin-bottom: 28px; font-size: 14px; }
  .charts {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 20px;
  }
  .chart-container {
    background: white;
    border-radius: 10px;
    padding: 20px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08);
    display: flex;
    flex-direction: column;
  }
  .chart-container h2 {
    font-size: 13px;
    color: #4f46e5;
    margin: 0 0 4px 0;
    font-weight: 600;
    letter-spacing: 0.5px;
    text-transform: uppercase;
  }
  .chart-container .meta {
    font-size: 12px;
    color: #9ca3af;
    margin-bottom: 12px;
  }
  .chart-wrapper { height: 380px; position: relative; }
  .prompt {
    font-style: italic;
    color: #374151;
    margin-top: 14px;
    padding: 10px 12px;
    background: #f3f4f6;
    border-left: 3px solid #4f46e5;
    border-radius: 4px;
    font-size: 13px;
    line-height: 1.4;
  }
  .prompt .ellipsis { color: #9ca3af; }
  footer {
    margin-top: 32px;
    padding-top: 16px;
    border-top: 1px solid #e5e7eb;
    font-size: 12px;
    color: #9ca3af;
  }
  @media (max-width: 1100px) {
    .charts { grid-template-columns: 1fr; }
  }
</style>
</head>
<body>
<h1>Distribuição de respostas do Claude Haiku 4.5 — efeito do contexto</h1>
<p class="subtitle">
  1000 chamadas por prompt · temperatura 1.0 · prefill como mensagem do assistant ·
  primeira palavra extraída de cada resposta · top 10 mostrado.
</p>
<div class="charts">
  <div class="chart-container">
    <h2>Q1 — Pouco contexto</h2>
    <div class="meta" id="meta1"></div>
    <div class="chart-wrapper"><canvas id="chart1"></canvas></div>
    <div class="prompt" id="prompt1"></div>
  </div>
  <div class="chart-container">
    <h2>Q2 — Contexto médio</h2>
    <div class="meta" id="meta2"></div>
    <div class="chart-wrapper"><canvas id="chart2"></canvas></div>
    <div class="prompt" id="prompt2"></div>
  </div>
  <div class="chart-container">
    <h2>Q3 — Muito contexto</h2>
    <div class="meta" id="meta3"></div>
    <div class="chart-wrapper"><canvas id="chart3"></canvas></div>
    <div class="prompt" id="prompt3"></div>
  </div>
</div>
<footer>
  Gerado por <code>autocomplete_simulation.py</code>. Aula de LLMs para advogados.
</footer>
<script>
const DATA = __DATA__;

function setPrompt(id, text) {
  const el = document.getElementById(id);
  el.innerHTML = '"' + text + '<span class="ellipsis">…</span>"';
}

function setMeta(id, info) {
  const el = document.getElementById(id);
  let msg = info.total_valid + ' respostas válidas';
  if (info.errors > 0) msg += ' · ' + info.errors + ' erros';
  el.textContent = msg;
}

function makeChart(canvasId, key) {
  const d = DATA[key];
  const ctx = document.getElementById(canvasId);
  new Chart(ctx, {
    type: 'bar',
    data: {
      labels: d.labels,
      datasets: [{
        label: '%',
        data: d.percentages,
        backgroundColor: 'rgba(79, 70, 229, 0.85)',
        borderColor: 'rgba(79, 70, 229, 1)',
        borderWidth: 1,
      }],
    },
    options: {
      indexAxis: 'y',
      responsive: true,
      maintainAspectRatio: false,
      plugins: {
        legend: { display: false },
        tooltip: {
          callbacks: {
            label: function(c) {
              return c.parsed.x.toFixed(1) + '%  (' + d.counts[c.dataIndex] + ' de ' + d.total_valid + ')';
            }
          }
        }
      },
      scales: {
        x: {
          title: { display: true, text: '% das respostas' },
          ticks: { callback: function(v) { return v + '%'; } }
        },
        y: {
          ticks: { font: { size: 13 } }
        }
      }
    }
  });
}

['Q1_pouco_contexto', 'Q2_contexto_medio', 'Q3_muito_contexto'].forEach((key, i) => {
  const n = i + 1;
  setPrompt('prompt' + n, DATA[key].prompt);
  setMeta('meta' + n, DATA[key]);
  makeChart('chart' + n, key);
});
</script>
</body>
</html>
""".replace("__DATA__", data_json)


async def main():
    print("=" * 70)
    print(f"Simulação de autocomplete — {MODEL_DISPLAY_NAME}")
    print("=" * 70)
    api_key = get_api_key()
    client = AsyncAnthropic(api_key=api_key)

    model = await detect_model(client)

    total = len(QUESTIONS) * N_PER_QUESTION
    print(f"\nTotal de chamadas: {total} ({N_PER_QUESTION} x {len(QUESTIONS)} perguntas)")
    print(f"Paralelismo: {CONCURRENCY} | Temperatura: {TEMPERATURE} | max_tokens: {MAX_TOKENS}")
    print(f"Custo estimado: ~US$ 0.20 (Haiku 4.5)\n")

    semaphore = asyncio.Semaphore(CONCURRENCY)
    progress = Progress(total)
    t0 = time.time()

    results = {}
    for q_key, prefill in QUESTIONS.items():
        print(f"\n[{q_key}] prefill: {prefill!r}")
        tasks = [
            call_one(client, model, prefill, semaphore, progress)
            for _ in range(N_PER_QUESTION)
        ]
        responses = await asyncio.gather(*tasks)
        results[q_key] = {"prompt": prefill, "responses": responses}

    elapsed = time.time() - t0
    print(f"\nConcluído em {elapsed:.1f}s\n")

    # Salvar bruto
    with open("raw_results.json", "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print("Salvo: raw_results.json")

    distributions = build_distributions(results)

    write_excel(distributions, "distribuicao.xlsx")
    print("Salvo: distribuicao.xlsx")

    html = build_html(distributions)
    with open("distribuicao.html", "w", encoding="utf-8") as f:
        f.write(html)
    print("Salvo: distribuicao.html")

    print("\n" + "=" * 70)
    print("TOP 5 por pergunta")
    print("=" * 70)
    for q_key, info in distributions.items():
        prefill = QUESTIONS[q_key]
        counter = info["counter"]
        total_valid = info["total_valid"]
        print(f"\n[{q_key}]  {prefill!r}")
        print(f"  válidas: {total_valid} · erros: {info['errors']}")
        for word, count in counter.most_common(5):
            pct = 100 * count / total_valid if total_valid else 0
            bar = "█" * int(pct / 2)
            print(f"  {word:<25} {count:>4}  ({pct:5.1f}%) {bar}")
    print()


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        sys.stderr.write("\nInterrompido pelo usuário.\n")
        sys.exit(130)
